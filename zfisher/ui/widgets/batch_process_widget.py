from magicgui.widgets import Container, PushButton, FileEdit, Label
from pathlib import Path
import logging
import napari
import pandas as pd
from qtpy.QtWidgets import QFrame, QFileDialog

from ...core import pipeline
from ...core.generate_batch_template import (
    build_template_sheets,
    add_dropdown_validations,
    parse_batch_config,
    validate_channels,
    resolve_puncta_for_dataset,
    resolve_coloc_for_dataset,
)
from .. import popups
from ..decorators import error_handler
from ..style import COLORS

logger = logging.getLogger(__name__)


class BatchProcessWidget(Container):
    def __init__(self, viewer: napari.Viewer):
        super().__init__(labels=False)
        self._viewer = viewer

        self._init_widgets()
        self._init_layout()
        self._connect_signals()

    def _init_widgets(self):
        self._header = Label(value="Batch Process")
        self._header.native.setObjectName("widgetHeader")
        self._info = Label(
            value="<i>Run the complete pipeline on multiple datasets using a batch Excel file.</i>"
        )
        self._info.native.setObjectName("widgetInfo")
        self._info.native.setWordWrap(True)
        from qtpy.QtWidgets import QSizePolicy
        self._info.native.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._batch_file = FileEdit(
            label="Batch File:",
            filter="*.xlsx *.xls",
            tooltip="Excel file (.xlsx) containing batch configuration. Use Generate Template to create one."
        )
        self._batch_output_dir = FileEdit(
            label="Output:",
            mode='d',
            value=Path.home() / "zFISHer_Batch_Output",
            tooltip="Base output directory. Each dataset gets its own subfolder."
        )
        self._generate_template_btn = PushButton(text="Generate Template")
        self._generate_template_btn.tooltip = "Create a template Excel file to configure batch processing with dataset paths and analysis parameters."
        self._generate_template_btn.native.setMinimumWidth(0)
        self._batch_run_btn = PushButton(text="Run Batch Processing")
        self._batch_run_btn.tooltip = "Validate the batch file and run the full pipeline on all datasets."
        self._batch_run_btn.native.setMinimumWidth(0)

    @staticmethod
    def _make_divider():
        line = QFrame()
        line.setFixedHeight(2)
        line.setStyleSheet(f"background-color: {COLORS['separator_color']}; border: none; margin: 8px 0px;")
        return line

    def _init_layout(self):
        from qtpy.QtWidgets import QLabel, QSpacerItem, QSizePolicy

        self._batch_form = Container(labels=True)
        self._batch_form.extend([self._batch_file])
        self._batch_form.native.layout().setContentsMargins(0, 10, 0, 0)

        self._out_form = Container(labels=True)
        self._out_form.extend([self._batch_output_dir])
        self._out_form.native.layout().setContentsMargins(0, 10, 0, 20)

        self._desc = QLabel(
            "Use Generate Template to create a pre-formatted workbook. "
            "Fill out the configuration in the workbook with Datasets, Puncta, and Colocalization sheets. "
            "Load the completed workbook and run to automate processing of all datasets defined in the "
            "workbook according to the configuration specified in each sheet."
        )
        self._desc.setWordWrap(True)
        self._desc.setStyleSheet("color: white; margin: 4px 2px;")

        spacer = lambda: QSpacerItem(0, 10, QSizePolicy.Minimum, QSizePolicy.Fixed)

        _layout = self.native.layout()
        _layout.setSpacing(2)
        _layout.setContentsMargins(0, 0, 0, 0)
        _layout.addWidget(self._header.native)
        _layout.addWidget(self._info.native)
        _layout.addWidget(self._make_divider())
        _layout.addWidget(self._desc)
        _layout.addWidget(self._batch_form.native)
        _layout.addWidget(self._out_form.native)
        _layout.addWidget(self._generate_template_btn.native)
        _layout.addSpacerItem(spacer())
        _layout.addWidget(self._batch_run_btn.native)
        _layout.addStretch(1)

    def _connect_signals(self):
        self._generate_template_btn.clicked.connect(self._on_generate_template)
        self._batch_run_btn.clicked.connect(self._on_batch_run)

    def _on_generate_template(self):
        """Save a multi-sheet batch template Excel file to a user-chosen location."""
        save_path, _ = QFileDialog.getSaveFileName(
            self.native,
            "Save Batch Template",
            str(Path.home() / "zFISHer_batch_template.xlsx"),
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return

        sheets = build_template_sheets()
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)
            add_dropdown_validations(writer.book)

            # --- Visual formatting ---
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Color palette
            header_fill = PatternFill(start_color="2D2B3D", end_color="2D2B3D", fill_type="solid")  # Dark purple
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            example_fill = PatternFill(start_color="3A3650", end_color="3A3650", fill_type="solid")  # Muted purple
            example_font = Font(name="Calibri", size=10, color="9999AA", italic=True)
            data_font = Font(name="Calibri", size=10, color="333333")
            section_fill = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")  # Light lavender
            section_font = Font(name="Calibri", size=11, bold=True, color="4A3A6A")
            desc_font = Font(name="Calibri", size=10, color="555555")
            thin_border = Border(
                bottom=Side(style="thin", color="CCCCCC"),
            )

            # --- Instructions sheet ---
            ws = writer.book["Instructions"]
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 120
            ws.sheet_properties.tabColor = "7A6B8A"

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                a_val = str(row[0].value or "")
                if a_val.isupper() and a_val.strip():
                    # Section headers (OVERVIEW, DATASETS SHEET, etc.)
                    for cell in row:
                        cell.fill = section_fill
                        cell.font = section_font
                elif a_val.startswith("---"):
                    # Sub-section separators
                    for cell in row:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="7A6B8A")
                elif a_val.startswith("("):
                    # Subtitle rows like (required), (optional)
                    for cell in row:
                        cell.font = Font(name="Calibri", size=10, italic=True, color="7A6B8A")
                elif a_val.strip():
                    # Column name rows
                    row[0].font = Font(name="Calibri", size=10, bold=True, color="333333")
                    if len(row) > 1:
                        row[1].font = desc_font
                        row[1].alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    # Description continuation rows
                    if len(row) > 1 and row[1].value:
                        row[1].font = desc_font
                        row[1].alignment = Alignment(wrap_text=True, vertical="top")

            # --- Data sheets (Datasets, Puncta, Colocalization) ---
            for sheet_name in ["Datasets", "Puncta", "Colocalization"]:
                ws = writer.book[sheet_name]
                ws.sheet_properties.tabColor = "7A6B8A"

                # Header row styling
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 28

                # Freeze top row
                ws.freeze_panes = "A2"

                # Example row styling
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    first_val = str(row[0].value or "")
                    if first_val.startswith("[EXAMPLE]"):
                        for cell in row:
                            cell.fill = example_fill
                            cell.font = example_font
                    else:
                        for cell in row:
                            cell.font = data_font
                            cell.border = thin_border

                # Auto-size columns based on header + a few data rows
                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    for row_idx in range(1, min(6, ws.max_row + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = max(14, min(40, max_len + 4))

        logger.info("Batch template saved to %s", save_path)
        popups.show_info_popup(
            self.native,
            "Template Saved",
            f"Batch template saved to:\n{save_path}"
        )

    # ------------------------------------------------------------------
    # Run
    # ------------------------------------------------------------------
    @error_handler("Batch Processing Failed")
    def _on_batch_run(self):
        excel_path = Path(self._batch_file.value)
        output_base = Path(self._batch_output_dir.value)

        if not excel_path.is_file():
            popups.show_error_popup(
                self._viewer.window._qt_window,
                "No Excel File",
                "Please select a valid batch Excel file (.xlsx)."
            )
            return

        if not output_base:
            popups.show_error_popup(
                self._viewer.window._qt_window,
                "No Output Directory",
                "Please select an output directory."
            )
            return

        # Parse and validate all sheets
        config, error = parse_batch_config(excel_path)
        if error:
            popups.show_error_popup(
                self._viewer.window._qt_window,
                "Batch Validation Failed",
                error
            )
            return

        # Collect all warnings: parse warnings + channel validation warnings
        all_warnings = config.pop("_warnings", [])

        # Deep validation: check file channels match config
        channel_warnings = validate_channels(config)
        all_warnings.extend(channel_warnings)

        if all_warnings:
            msg = (
                "The following issues were found during validation:\n\n"
                + "\n".join(f"  - {w}" for w in all_warnings)
                + "\n\nDo you want to continue anyway?"
            )
            if not popups.show_yes_no_popup(
                self._viewer.window._qt_window,
                "Batch Validation Warnings",
                msg,
            ):
                return
        else:
            # No warnings — confirm start
            if not popups.show_yes_no_popup(
                self._viewer.window._qt_window,
                "Batch Validation Passed",
                f"All {len(config['datasets'])} dataset(s) validated successfully.\n\n"
                "Start batch processing?",
            ):
                return

        output_base.mkdir(parents=True, exist_ok=True)

        num_items = len(config["datasets"])
        results = []

        with popups.BatchProgressDialog(
            self._viewer.window._qt_window,
            title="Batch Processing",
            text="Preparing..."
        ) as dialog:
            for i, ds in enumerate(config["datasets"]):
                name = ds["name"]
                item_output = ds["output_dir"] or (output_base / name)

                batch_pct = int((i / num_items) * 100)
                dialog.update_batch_progress(
                    batch_pct,
                    f"Batch: {i + 1} / {num_items} — {name}"
                )
                dialog.update_item_progress(0, f"Starting {name}...")

                # Resolve per-dataset puncta and colocalization config
                puncta_config = resolve_puncta_for_dataset(name, config["puncta_rules"])
                pairwise_rules, tri_rules = resolve_coloc_for_dataset(name, config["coloc_rules"])

                try:
                    pipeline.run_full_zfisher_pipeline(
                        r1_path=ds["r1"],
                        r2_path=ds["r2"],
                        output_dir=item_output,
                        seg_method=ds["seg_method"],
                        merge_splits=ds["merge_splits"],
                        r1_nuclear_channel=ds["r1_nuclear_channel"],
                        r2_nuclear_channel=ds["r2_nuclear_channel"],
                        puncta_config=puncta_config,
                        pairwise_rules=pairwise_rules,
                        tri_rules=tri_rules,
                        apply_warp=ds.get("apply_warp", True),
                        max_ransac_distance=ds.get("max_ransac_distance", 0),
                        overlap_method=ds.get("overlap_method", "Intersection"),
                        match_threshold=ds.get("match_threshold", 0),
                        remove_extranuclear_puncta=ds.get("remove_extranuclear_puncta", True),
                        progress_callback=dialog.update_item_progress,
                    )
                    results.append((name, "Success"))
                    logger.info("Batch item '%s' completed successfully.", name)
                except Exception as exc:
                    logger.error("Batch item '%s' failed: %s", name, exc, exc_info=True)
                    results.append((name, f"Failed: {exc}"))

            dialog.update_batch_progress(100, "Batch complete.")

        # Summary popup
        lines = [f"  {name}: {status}" for name, status in results]
        summary = "\n".join(lines)
        popups.show_info_popup(
            self._viewer.window._qt_window,
            "Batch Processing Complete",
            f"Processed {num_items} item(s):\n\n{summary}"
        )
